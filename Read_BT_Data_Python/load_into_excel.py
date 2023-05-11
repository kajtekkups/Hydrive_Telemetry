import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook

# filename = sys.argv[1]
#nazwa, rozszerzenie = os.path.splitext(FileName)
# excelName = nazwa


czas = []
pomiarVT_0 = []
pomiarVT_1 = []
pomiarI_0 = []
pomiarI_1 = []

czasSlowo = 'czas:'
VT_0Slowo = 'pomiarVT_0:'
VT_1Slowo = 'pomiarVT_1:'

I_0Slowo = 'pomiarI_0:'
I_1Slowo = 'pomiarI_1:'

# tabele = [czas, pomiarVT_0, pomiarVT_1, pomiarI_0, pomiarI_1]
# slowaKlucz = [czasSlowo, VT_0Slowo, VT_1Slowo, I_0Slowo, I_1Slowo]



pomiarVT_2 = []
pomiarI_2 = []
VT_2Slowo = 'pomiarVT_2:'
I_2Slowo = 'pomiarI_2:'
tabele = [czas, pomiarVT_0, pomiarVT_1, pomiarI_0, pomiarI_1, pomiarVT_2, pomiarI_2]
slowaKlucz = [czasSlowo, VT_0Slowo, VT_1Slowo, I_0Slowo, I_1Slowo, VT_2Slowo, I_2Slowo]




def ReadValue(tekst, slowoKlucz):
    tekst = tekst.replace(" ", "")

    firstIndex = tekst.rfind(slowoKlucz) + len(slowoKlucz)

    digitsNumber = 0
    while tekst[digitsNumber + firstIndex] != '[':
        digitsNumber +=1

    value = tekst[firstIndex:firstIndex + digitsNumber]

    return float(value)

def saveInExcel(columnName, columnNumber, data):

    try:
        excel = load_workbook(filename='HydrivePomiary.xlsx')
    except:
        # Tworzenie nowego pliku Excela
        excel = Workbook()

    # Tworzenie nowego arkusza
    worksheet = excel.active

    # Ustawienie nagłówka dla kolumny A
    worksheet.cell(row=1, column=columnNumber, value=columnName)

    # Dodanie danych do kolumny, począwszy od wiersza 2
    for i, value in enumerate(data, start=2):
        worksheet.cell(row=i, column=columnNumber, value=value)

    # Zapisanie pliku
    excel.save('HydrivePomiary.xlsx')

def ColectData():

    for number, line in enumerate(lines):
        if not(100 < len(line) < 250):
            continue

        line.strip()

        for index in range(len(tabele)):
            tabele[index].append(ReadValue(line, slowaKlucz[index]))

    for index in range(len(tabele)):
        saveInExcel(slowaKlucz[index], index+1, tabele[index])

# slowaKlucz = [czasSlowo, VT_0Slowo, VT_1Slowo, I_0Slowo, I_1Slowo]

lines = open('Pomiary/HydrivePomiaryPradu_0.txt', 'r').readlines()
ColectData()

