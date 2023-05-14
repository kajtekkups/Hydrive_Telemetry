# import threading
#
# def google_thread():
#     # kod dla wątku zapisującego dane do chmury Google
#     pass
#
# def excel_thread():
#     # kod dla wątku zapisującego dane do pliku Excel
#     pass
#
# def on_message(client, userdata, message):
#     # Kod do odczytywania i przetwarzania danych z serwera MQTT
#     # Gdy dane zostaną odczytane, można uruchomić wątki zapisujące do Google i Excel
#     google_t = threading.Thread(target=google_thread)
#     excel_t = threading.Thread(target=excel_thread)
#     google_t.start()
#     excel_t.start()
#
# # Kod do subskrybowania kanału MQTT i odbierania danych
# client = mqtt.Client()
# client.on_message = on_message
# client.connect("broker.hivemq.com", 1883, 60)
# client.subscribe("test/topic")
# client.loop_forever()

from AWSIoTPythonSDK.MQTTLib import AWSIoTMQTTClient
import json
import sys
import threading
import time
from openpyxl import Workbook
from openpyxl import load_workbook
import os

#Ustawienia serwera aws
myMQTTClient = AWSIoTMQTTClient("Hydrive_komputer")
myMQTTClient.configureEndpoint("andh08iyq0pvi-ats.iot.eu-north-1.amazonaws.com", 443)
myMQTTClient.configureCredentials("./certyfikaty_komp/AmazonRootCA1.pem", "./certyfikaty_komp/private.pem.key", "./certyfikaty_komp/certificate.pem.crt.txt")
myMQTTClient.connect()
print("Client Connected")
topic = "Hydrive_komputer_sub"

#lista wartosci z ostatnich x wynikow
ostatnie_pomiary = []
# plik Excela
EXCEL_FILE = "nazwa_pliku.xlsx"
EXCEL_SHEET = "nazwa_arkusza"

ile_danych_w_danym_cyklu = 0
data_to_google = {
    "time": 0,
    "pomiarVT_0": 0,
    "pomiarVT_1": 0,
    "pomiarV_2": 0,
    "pomiarI_0": 0,
    "pomiarI_1": 0,
    "pomiarI_2": 0
}

suma_wartosci_zmierzonych = [0 for w in range(7)]

# kod dla wątku zapisującego dane do pliku Excel
def saveInExcel(data):


    try:
        excel = load_workbook(filename='HydrivePomiary.xlsx')
    except:
        # Tworzenie nowego pliku Excela
        excel = Workbook()

    # Tworzenie nowego arkusza
    worksheet = excel.active

    # Znajdź ostatni wiersz z danymi
    last_row = worksheet.max_row + 1

    # Dodanie danych do kolumny, począwszy od wiersza 2
    for i, value in enumerate(data):
        worksheet.cell(row=last_row, column=(i + 1), value=value)

    # Zapisanie pliku
    excel.save('HydrivePomiary.xlsx')


# kod dla wątku zapisującego dane do chmury Google
def google_thread(data):
    global data_to_google
    global ile_danych_w_danym_cyklu
    global suma_wartosci_zmierzonych

    if os.path.isfile('data.json'):
        ile_danych_w_danym_cyklu += 1
        # licz srednia z ile_danych_w_danym_cyklu
        data_to_google["time"] = data[0]
        data_to_google["pomiarVT_0"] = (suma_wartosci_zmierzonych[1] / ile_danych_w_danym_cyklu)
        # data_to_google = {
        #     "time": 0,
        #     "pomiarVT_0": 0,
        #     "pomiarVT_1": 0,
        #     "pomiarV_2": 0,
        #     "pomiarI_0": 0,
        #     "pomiarI_1": 0,
        #     "pomiarI_2": 0
        # }

    else:
        with open("data.json", "w") as f:
            json.dump(data_to_google, f)

        #zresetuj dane
        data_to_google = {
            "time": 0,
            "pomiarVT_0": 0,
            "pomiarVT_1": 0,
            "pomiarV_2": 0,
            "pomiarI_0": 0,
            "pomiarI_1": 0,
            "pomiarI_2": 0
        }
        suma_wartosci_zmierzonych = [0 for i in range(7)]

        ile_danych_w_danym_cyklu = 0



def handle_message(client, userdata, message):

    # convert message payload to character buffer
    json_str = message.payload.decode('utf-8')

    # deserialize JSON
    json_obj = json.loads(json_str)

    # access JSON data
    data = [json_obj["time"], json_obj["pomiarVT_0"], json_obj["pomiarVT_1"], json_obj["pomiarV_2"], json_obj["pomiarI_0"],
            json_obj["pomiarI_1"], json_obj["pomiarI_2"]]

    print(f"Received message:  on topic {message.topic}")
    print(data[0])

    # dekodowanie i przetwarzanie danych z MQTT
    # wywołanie funkcji dla każdego wątku
    # saveInExcel(data)
    google_thread(data)


myMQTTClient.subscribe(topic, 0, handle_message)
print("Message Sent")

while True:
    pass